from openpyxl import load_workbook
from model.Student_class import Student

"""
                   File Function Description
scores_divided[used in sort]        -return a list with all the students scores and sum
sort                  -sort by subject

change score          -change a specific student's specific score
insert                -insert a student's information
part_operate          -use in a terminal
"""


def scores_divided(students):
    list_Chinese = []
    list_Math = []
    list_English = []
    list_Total = []
    for student in students:
        list_Chinese.append(student.list_scores[0])
        list_Math.append(student.list_scores[1])
        list_English.append(student.list_scores[2])
        list_Total.append(student.sum())
    list = [list_Chinese, list_Math, list_English, list_Total]
    return list


def sort(students, choose00, subjectChoose):
    choose0 = choose00
    def bubble(students, list):  # bubble sort
        choose_1 = 0
        try:
            choose_1 = int(choose0)
        except:
            print('')
        if choose_1 == 1:
            for i in range(1, len(list)):
                for j in range(0, len(list) - i):
                    if list[j] > list[j + 1]:
                        students[j], students[j + 1] = students[j + 1], students[j]
                        list[j], list[j + 1] = list[j + 1], list[j]
        elif choose_1 == 2:
            for i in range(1, len(list)):
                for j in range(0, len(list) - i):
                    if list[j] < list[j + 1]:
                        students[j], students[j + 1] = students[j + 1], students[j]
                        list[j], list[j + 1] = list[j + 1], list[j]
        else:
            print('invalid input')

    def selection(students, list):  # choose sort
        choose_1 = 0
        try:
            choose_1 = int(choose0)
        except:
            print('')
        if choose_1 == 1:
            for i in range(0, len(list)):
                num = i
                for j in range(i + 1, len(list)):
                    if list[num] > list[j]:
                        num = j
                if num != i:
                    students[i], students[num] = students[num], students[i]
                    list[i], list[num] = list[num], list[i]
        elif choose_1 == 2:
            for i in range(0, len(list)):
                num = i
                for j in range(i + 1, len(list)):
                    if list[num] < list[j]:
                        num = j
                if num != i:
                    students[i], students[num] = students[num], students[i]
                    list[i], list[num] = list[num], list[i]
        else:
            print('invalid input')

    choose = 0
    try:
        choose = int(subjectChoose)
    except:
        print('')
    if choose == 1:
        list_scores_all = scores_divided(students)
        bubble(students, list_scores_all[0])
    elif choose == 2:
        list_scores_all = scores_divided(students)
        selection(students, list_scores_all[1])
    elif choose == 3:
        list_scores_all = scores_divided(students)
        bubble(students, list_scores_all[2])
    elif choose == 4:
        list_scores_all = scores_divided(students)
        selection(students, list_scores_all[3])
    else:
        print('invalid input')


def change_score(students, nameId, subject, newScore):
    for student in students:
        if student.name == nameId or str(student.id) == nameId:
            subject_ = int(subject)
            if subject_ == 1:
                try:
                    new_score = int(newScore)
                    student.list_scores[0] = new_score
                except:
                    print('invalid input')
            elif subject_ == 2:
                try:
                    new_score = int(newScore)
                    student.list_scores[1] = new_score
                except:
                    print('invalid input')
            elif subject_ == 3:
                try:
                    new_score = int(newScore)
                    student.list_scores[2] = new_score
                except:
                    print('invalid input')
            else:
                print('invalid input')


def insert(students, newName, newId, newChinese, newMath, newEnglish):
    name = newName
    while True:
        try:
            id = int(newId)
            Chinese = float(newChinese)
            Math = float(newMath)
            English = float(newEnglish)
            break
        except:
            print('invalid input,please try again！')
    new_student = Student(name, id, Chinese, Math, English)
    students.append(new_student)


def print_student_info(students):
    for student in students:
        print(f"Student Name: {student.name}")
        print(f"Student ID:   {student.id}")
        print(
            f"Chinese: {student.list_scores[0]},   Math: {student.list_scores[1]},   English: {student.list_scores[2]}")
        print(
            f"Grades:  {student.list_grades[0]}             {student.list_grades[1]}                {student.list_grades[2]}")
        print("---------------------------------------------")



def xlsx_update(students):
    wb = load_workbook(filename='Scores_Original.xlsx', read_only=False)
    Sheet = wb['Sheet1']
    count = 0
    for row in range(2, Sheet.max_row+1):
        Sheet.cell(row,1).value = students[count].name
        Sheet.cell(row,2).value = students[count].id
        Sheet.cell(row,3).value = students[count].list_scores[0]
        Sheet.cell(row,4).value = students[count].list_scores[1]
        Sheet.cell(row,5).value = students[count].list_scores[2]
        count = count+1
    wb.save('Scores_Original.xlsx')